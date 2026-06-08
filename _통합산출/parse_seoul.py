# -*- coding: utf-8 -*-
"""서울 공무원 xlsx(전체 시트) -> 통일 양식 CSV. 원본은 읽기 전용."""
import openpyxl, glob, os, re, csv

ROOT = r"C:/Users/admin/Documents/이민범 개발/합격선/합격선 최종모음"
OUT = r"C:/Users/admin/Documents/이민범 개발/합격선/_통합산출/합격선_통합.csv"

def folder_meta(folder):
    """'01 서울 공무원' -> ('서울','공무원')"""
    name = re.sub(r"^\d+\s*", "", folder).strip()
    parts = name.split()
    region = parts[0] if parts else name
    kind = " ".join(parts[1:]) if len(parts) > 1 else ""
    return region, kind

HEADERS = ["지역","시험종류","연도","회차","시험구분","임용예정기관","직군","직렬","직류","직급","대상",
           "선발예정인원","접수인원","접수여성","경쟁률(접수/선발)","응시인원","응시여성","응시율","경쟁률(응시/선발)",
           "필기합격인원","필기합격여성","합격선","양성평등성별","양성평등합격선","최종합격인원","최종합격여성",
           "원본파일명","비고"]

def expand_merges(ws):
    """병합셀을 풀어 좌상단 값을 범위 전체에 채운 2차원 리스트 반환."""
    maxr, maxc = ws.max_row, ws.max_column
    grid = [[ws.cell(r, c).value for c in range(1, maxc+1)] for r in range(1, maxr+1)]
    for rng in ws.merged_cells.ranges:
        tl = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row+1):
            for c in range(rng.min_col, rng.max_col+1):
                grid[r-1][c-1] = tl
    return grid

def is_empty(v):
    return v is None or (isinstance(v, str) and v.strip() == "")

def num_or_blank(v):
    if is_empty(v): return ""
    return v

def parse_filename(fn):
    yr = re.search(r"(\d{4})년도", fn)
    rd = re.search(r"제\s*(\d+)\s*회", fn)
    return (yr.group(1) if yr else ""), (rd.group(1)+"회" if rd else "")

def parse_jikryu(c2):
    m = re.match(r"^(.*?)\((.*?)\)\s*$", str(c2).strip())
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return str(c2).strip(), "일반"

def pct(v):
    """0~1 분수면 %로 변환(소수1자리)."""
    if is_empty(v): return ""
    try:
        f = float(v)
        if f <= 1.0: return round(f*100, 1)
        return round(f, 1)
    except: return v

rows_out = []
skipped = []
xlsx_files = sorted(glob.glob(os.path.join(ROOT, "*", "*.xlsx")))
for f in xlsx_files:
    fn = os.path.basename(f)
    region, kind = folder_meta(os.path.basename(os.path.dirname(f)))
    yr, rd = parse_filename(fn)
    wb = openpyxl.load_workbook(f, data_only=True)
    if "전체" not in wb.sheetnames:
        skipped.append((fn, "전체 시트 없음(서울 양식 아님)"))
        continue
    ws = wb["전체"]
    grid = expand_merges(ws)
    cur_gubun = cur_group = cur_jikryeol = ""
    cnt = 0
    for row in grid:
        c = (list(row) + [None]*19)[:19]
        c0,c1,c2,c3 = c[0],c[1],c[2],c[3]
        if all(is_empty(x) for x in c[:4]): continue
        s0 = str(c0).strip() if not is_empty(c0) else ""
        # 시험구분 (공개경쟁/경력경쟁/특별...)
        if s0 and s0 not in ("구분",) and not s0.startswith("총계"):
            cur_gubun = s0
        if s0.startswith("총계"): continue
        # 그룹/직군 행 (직급 없음)
        if is_empty(c3):
            if not is_empty(c1) and str(c1).strip() != "소계":
                lab = str(c1).strip()
                cur_group = lab
                cur_jikryeol = ""
            continue
        # ----- leaf 행 (직급 존재) -----
        # 데이터 행은 선발예정인원(c4)이 숫자. 제목/헤더 행 제거.
        if not isinstance(c[4], (int, float)):
            continue
        if not is_empty(c1) and str(c1).strip() != "소계":
            cur_jikryeol = str(c1).strip()
        jikryu, daesang = parse_jikryu(c2)
        # 양성평등
        gnd = c[15]; gcut = c[16]
        if is_empty(gnd) or str(gnd).strip() in ("0","0.0"):
            gnd, gcut = "", ""
        cut = c[14]
        if is_empty(cut) or str(cut).strip() in ("0","0.0"): cut = ""
        rows_out.append([
            region,kind,yr,rd,cur_gubun,"",cur_group,cur_jikryeol,jikryu,str(c3).strip(),daesang,
            num_or_blank(c[4]),num_or_blank(c[5]),num_or_blank(c[6]),num_or_blank(c[7]),
            num_or_blank(c[8]),num_or_blank(c[9]),pct(c[10]),num_or_blank(c[11]),
            num_or_blank(c[12]),num_or_blank(c[13]),cut,gnd,gcut,
            num_or_blank(c[17]),num_or_blank(c[18]),fn,""
        ])
        cnt += 1
    print(f"{fn} -> {cnt} rows")

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8-sig", newline="") as fp:
    w = csv.writer(fp)
    w.writerow(HEADERS)
    w.writerows(rows_out)
print("TOTAL:", len(rows_out), "->", OUT)
if skipped:
    print("--- skipped (xlsx, 비서울양식) ---")
    for s in skipped: print("  ", s)
print("--- sample ---")
for r in rows_out[:4] + rows_out[-3:]:
    print(r)
