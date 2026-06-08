# -*- coding: utf-8 -*-
"""4시트 xlsx: 1)작업메모 2)합격선통합(공개경쟁) 3)2026접수현황 4)합격선_경력경쟁.
계산 가능한 경쟁률/응시율 채움. 시트1은 미반영파일·접수현황없음·확인필요 자동수집."""
import csv, os, glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

D = os.path.dirname(__file__)
MAIN = os.path.join(D,"합격선_통합_merged.csv")
JEOP = os.path.join(D,"접수현황_data.csv")
OUT  = os.path.join(D,"합격선_통합.xlsx")
SRC  = r"C:/Users/admin/Documents/이민범 개발/합격선/합격선 최종모음"

NUMCOLS = {"선발예정인원","접수인원","접수여성","응시인원","응시여성","응시율",
           "필기합격인원","필기합격여성","합격선","양성평등합격선","최종합격인원","최종합격여성"}

def to_num(v):
    if v is None or v=="": return ""
    try:
        f=float(v); return int(f) if f==int(f) else f
    except: return v

def fnum(v):
    try: return float(str(v).replace(',',''))
    except: return None

def load(path):
    with open(path, encoding="utf-8-sig") as fp:
        rows=list(csv.reader(fp))
    return rows[0], rows[1:]

# ---------- MAIN 로드 + 합격선기준 + 경쟁률/응시율 계산 ----------
header, data = load(MAIN)
pos=header.index("양성평등합격선")+1
ki=header.index("시험종류"); ci=header.index("합격선")
header=header[:pos]+["합격선기준"]+header[pos:]
nd=[]
for r in data:
    cv=fnum(r[ci]) if 'fnum' in dir() else None
    try: cv=float(str(r[ci]).replace(',',''))
    except: cv=None
    if "교행" in (r[ki] or ""):
        std="총점" if (cv is not None and cv>130) else "과목평균"
    else:
        std="과목평균"
    if not (r[ci] or "").strip(): std=""
    nd.append(r[:pos]+[std]+r[pos:])
data=nd

I={h:i for i,h in enumerate(header)}
def compute(r):
    sel=fnum(r[I["선발예정인원"]]); jeop=fnum(r[I["접수인원"]]); eung=fnum(r[I["응시인원"]])
    if not r[I["경쟁률(접수/선발)"]].strip() and sel and jeop is not None and sel>0:
        r[I["경쟁률(접수/선발)"]]="%.1f:1"%(jeop/sel)
    if not r[I["응시율"]].strip() and jeop and eung is not None and jeop>0:
        r[I["응시율"]]=round(eung/jeop*100,1)
    if not r[I["경쟁률(응시/선발)"]].strip() and sel and eung is not None and sel>0:
        r[I["경쟁률(응시/선발)"]]="%.1f:1"%(eung/sel)
    return r
data=[compute(r) for r in data]

def _n(v):
    try: return float(str(v).replace(',',''))
    except: return None
def _fmt(v):
    return str(int(v)) if v==int(v) else str(round(v,2))
def aggregate(rows):
    from collections import OrderedDict
    keys=['지역','시험종류','연도','회차','시험구분','직군','직렬','직류','직급','대상']
    groups=OrderedDict()
    for r in rows:
        k=tuple(r[I[c]] for c in keys)
        groups.setdefault(k,[]).append(r)
    out=[]
    for k,grp in groups.items():
        a=['']*len(header)
        for i,c in enumerate(keys): a[I[c]]=k[i]
        def s(col):
            vs=[_n(r[I[col]]) for r in grp]; vs=[v for v in vs if v is not None]
            return _fmt(sum(vs)) if vs else ''
        for col in ['선발예정인원','접수인원','응시인원','필기합격인원','최종합격인원']:
            a[I[col]]=s(col)
        cuts=sorted(set(_n(r[I['합격선']]) for r in grp if _n(r[I['합격선']]) is not None))
        if cuts: a[I['합격선']]=_fmt(cuts[0]) if len(cuts)==1 else f"{_fmt(cuts[0])}~{_fmt(cuts[-1])}"
        yangs=sorted(set(_n(r[I['양성평등합격선']]) for r in grp if _n(r[I['양성평등합격선']]) is not None))
        if yangs: a[I['양성평등합격선']]=_fmt(yangs[0]) if len(yangs)==1 else f"{_fmt(yangs[0])}~{_fmt(yangs[-1])}"
        a[I['합격선기준']]=next((r[I['합격선기준']] for r in grp if r[I['합격선기준']]),'')
        gigs=sorted(set(r[I['임용예정기관']] for r in grp if r[I['임용예정기관']]))
        a[I['임용예정기관']]=(gigs[0] if len(gigs)==1 else (f"{len(gigs)}개 기관" if gigs else ''))
        sel,jeop,eung=_n(a[I['선발예정인원']]),_n(a[I['접수인원']]),_n(a[I['응시인원']])
        if jeop and sel: a[I['경쟁률(접수/선발)']]=f"{jeop/sel:.1f}:1"
        if eung and jeop: a[I['응시율']]=round(eung/jeop*100,1)
        if eung and sel: a[I['경쟁률(응시/선발)']]=f"{eung/sel:.1f}:1"
        a[I['원본파일명']]=f"{len(grp)}건 집계"
        notes=sorted(set(r[I['비고']] for r in grp if r[I['비고']]))
        a[I['비고']]=' / '.join(notes)[:60]
        out.append(a)
    return out
agg=aggregate(data)
gong=[r for r in data if "공개" in r[I["시험구분"]]]

# ---------- 시트1 작업메모 자동수집 ----------
todo=[]   # [유형,지역,시험종류,연도,회차,항목,상태,메모]
seq=0
def T(t,reg,kind,yr,rd,item,memo):
    global seq; seq+=1
    todo.append([seq,t,reg,kind,yr,rd,item,"",memo])

# (1) 미반영 파일: 최종모음 데이터파일 중 원본파일명에 없는 것
used=set(r[I["원본파일명"]] for r in data)
jh,jd=load(JEOP); ji={h:i for i,h in enumerate(jh)}
used|=set(r[ji["원본파일명"]] for r in jd)
# 보조 출처(합격선 행 보강에 사용한 응시/접수현황) — 미반영 제외
EXTRA_USED={
 "2023년 대구시교육청 지방공무원 필기시험 응시현황(공개).pdf",
 "2025년도 대구광역시교육청 지방공무원필기시험응시현황(공개).pdf",
 "2025+대구교육청+필기시험응시현황공개.pdf",
 "☆2024년+대구교육청+원서접수현황홈페이지+게시용.pdf",
 "☆2023년+원서접수현황홈페이지+게시용.pdf",
 # 서울·부산 교행 응시현황(합격선 행 보강에 사용)
 "2023년도 제1회 서울특별시교육청 지방공무원 9급 신규 임용 필기시험 응시 현황.pdf",
 "2024년도 서울특별시교육청 지방공무원 필기시험 응시율.pdf",
 "2025년도 서울시교육청 지방공무원 9급 공개(경력)경쟁임용 필기시험 응시현황(게시용).pdf",
 "2023년부산광역시교육청지방공무원임용필기시험응시현황.pdf",
 "2024년도 부산광역시교육청 지방공무원 필기시험 응시율.pdf",
 "2025년+부산광역시교육청+지방공무원+임용+필기시험+응시현황.pdf",
 "2025년도 부산광역시교육청 지방공무원 임용 필기시험 응시현황.pdf",
 "2024년도+부산광역시교육청+지방공무원+임용시험+응시원서+접수현황게시용.pdf",  # 2024응시현황으로 보강됨(중복)
 # 인천 교행 접수결과(접수 보강에 사용)
 "2023년도제2회인천광역시교육청지방공무원공개경력경쟁임용시험응시원서최종접수결과게시용.pdf",
 "2024년도+인천+응시원서+최종+접수+결과게시용.pdf",
 # 광주 교행 응시현황(선발/접수/응시 보강에 사용)
 "2023년 광주광역시교육청 지방공무원 필기시험 응시현황.pdf",
 "2024년도 광주광역시교육청 지방공무원 필기시험 응시율.pdf",
 "2025년도 광주광역시교육청 지방공무원 제1회 임용 필기시험 응시현황.pdf",
}
used|=EXTRA_USED
EXTS={".xlsx",".hwp",".hwpx",".pdf",".png",".csv",".xls"}
for f in sorted(glob.glob(os.path.join(SRC,"*","*"))):
    fn=os.path.basename(f); ext=os.path.splitext(fn)[1].lower()
    if ext not in EXTS: continue
    if fn in used: continue
    folder=os.path.basename(os.path.dirname(f))
    parts=folder.replace("0","").split()
    nm=folder
    import re as _re
    nm=_re.sub(r'^\d+\s*','',folder).split()
    reg=nm[0] if nm else folder; kind=" ".join(nm[1:]) if len(nm)>1 else ""
    T("미반영(읽기 필요)",reg,kind,"","",fn,"양식 확인 후 통합 필요(%s)"%ext)

# (2) 공개경쟁인데 접수현황(접수인원) 전무한 묶음
from collections import defaultdict
grp=defaultdict(lambda:[0,0])
for r in gong:
    k=(r[I["지역"]],r[I["시험종류"]],r[I["연도"]],r[I["회차"]])
    grp[k][0]+=1
    if r[I["접수인원"]].strip(): grp[k][1]+=1
for k in sorted(grp):
    if grp[k][1]==0:
        T("접수현황 찾기",k[0],k[1],k[2],k[3],"원서접수 인원(가능시 응시·여성)","현재 합격선만 보유 / 경력경쟁 제외")

# (3) 확인필요: 비고 플래그(직급확인/비공개)
seen=set()
for r in data:
    bg=r[I["비고"]]
    if not bg: continue
    if any(w in bg for w in ("직급확인","비공개","추정","미표기","확인")):
        k=(r[I["지역"]],r[I["시험종류"]],r[I["연도"]],r[I["회차"]],bg)
        if k in seen: continue
        seen.add(k)
        T("확인필요",k[0],k[1],k[2],k[3],bg,"원본 대조 권장")

# (4) 안내/수기
T("안내","","","","","범위: 전 지역 수집 확정. 새 파일은 자동으로 위 '미반영'에 표기됨","")
T("확인필요","인천","공무원","","","공채/경채 구분이 원본 통계자료에 없음 → 전부 공개경쟁으로 넣음","경채 직렬(간호·의료기술 등) 알려주면 시트4로 이동. 합격선=평균(총점은 비고)")

# ================= 시트 쓰기 =================
hfill=PatternFill("solid",fgColor="1F4E78"); hfont=Font(bold=True,color="FFFFFF",size=10)
t1fill=PatternFill("solid",fgColor="C55A11")
thin=Side(style="thin",color="D9D9D9"); border=Border(thin,thin,thin,thin)
WIDTHS={"지역":6,"시험종류":8,"연도":6,"회차":6,"시험구분":9,"임용예정기관":12,"직군":12,"직렬":11,
        "직류":13,"직급":7,"대상":10,"합격선":8,"합격선기준":10,"경쟁률(접수/선발)":13,
        "경쟁률(응시/선발)":13,"원본파일명":44,"비고":20,"항목":48,"메모":34,"유형":14,"상태":8}

def sheet(wb,title,hdr,rows,first=False,hcolor=hfill):
    ws=wb.active if first else wb.create_sheet(); ws.title=title
    numidx={i for i,h in enumerate(hdr) if h in NUMCOLS}
    ws.append(hdr)
    for r in rows:
        ws.append([to_num(v) if i in numidx else v for i,v in enumerate(r)])
    for c in range(1,len(hdr)+1):
        cell=ws.cell(1,c); cell.fill=hcolor; cell.font=hfont
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=border
    ws.freeze_panes="A2"
    ws.auto_filter.ref=f"A1:{ws.cell(1,len(hdr)).column_letter}{ws.max_row}"
    for i,h in enumerate(hdr,1):
        ws.column_dimensions[ws.cell(1,i).column_letter].width=WIDTHS.get(h,9)
    return ws.max_row-1

wb=Workbook()
H1=["번호","유형","지역","시험종류","연도","회차","항목","상태","메모"]
n1=sheet(wb,"작업메모",H1,todo,first=True,hcolor=t1fill)
n2=sheet(wb,"합격선_직렬별",header,agg)       # 집계(보기용)
n5=sheet(wb,"합격선_상세",header,data)         # 시·군별 원자료
jh2,jd2=load(JEOP)
jpos=jh2.index("양성평등합격선")+1
jh2=jh2[:jpos]+["합격선기준"]+jh2[jpos:]; jd2=[r[:jpos]+[""]+r[jpos:] for r in jd2]
n3=sheet(wb,"2026접수현황",jh2,jd2)
wb.save(OUT)
print(f"작업메모:{n1} / 직렬별집계:{n2} / 상세:{n5} / 2026접수:{n3}")
print("-> ",OUT)
