# -*- coding: utf-8 -*-
"""경기 공무원 xlsx(시·군별, 평균+양성) + 경기 교행 pdf(/묶음, 총득점) -> 통일 양식."""
import openpyxl, pdfplumber, re, os, glob, csv
ROOT=r"C:/Users/admin/Documents/이민범 개발/합격선/합격선 최종모음"
HEADERS=["지역","시험종류","연도","회차","시험구분","임용예정기관","직군","직렬","직류","직급","대상",
         "선발예정인원","접수인원","접수여성","경쟁률(접수/선발)","응시인원","응시여성","응시율","경쟁률(응시/선발)",
         "필기합격인원","필기합격여성","합격선","양성평등성별","양성평등합격선","최종합격인원","최종합격여성",
         "원본파일명","비고"]
def num(t):
    t=re.sub(r'[,\s명점]','',str(t or ''))
    return '' if t in ('','-','–','비공개') else t
def daenorm(s):
    s=re.sub(r'\s+','',str(s))
    if '저소득' in s: return '저소득층'
    if '장애' in s: return '장애인'
    if '보훈' in s: return '보훈청추천'
    if '지방의회' in s: return '지방의회'
    return s or '일반'
def yr_rd(fn):
    y=re.search(r'(\d{4})',fn); r=re.search(r'제\s*(\d+)\s*회',fn)
    return (y.group(1) if y else ''),(r.group(1)+'회' if r else '')
def parse_jik(s):  # '간호8급(간호)' / '행정9급(일반행정)'
    s=str(s).replace('\n','').strip()
    m=re.match(r'^(\D+?)(\d+급|연구사|지도사)\((.+)\)$', s)
    if m:
        jr=m.group(1).strip(); jg=m.group(2); inner=m.group(3)
        if '_' in inner: ju,dae=inner.rsplit('_',1); dae=daenorm(dae)
        else: ju,dae=inner,'일반'
    else:
        m2=re.match(r'^(.+?)\s*\((.+?)\)\s*(\d+급)?',s)
        if m2: jr,ju,jg=m2.group(1),m2.group(2),(m2.group(3) or '')
        else:
            m3=re.match(r'^(.+?)\s*(\d+급|연구사|지도사)$',s);
            if m3: jr=ju=m3.group(1); jg=m3.group(2)
            else: jr=ju=s; jg=''
        dae='일반'
    jr=re.sub(r'\s+','',jr)
    if jr.endswith('직') and len(jr)>1: jr=jr[:-1]
    return jr,re.sub(r'\s+','',ju),jg,dae

# ---------- 경기 공무원 xlsx ----------
def gong_xlsx(path):
    fn=os.path.basename(path); yr,rd=yr_rd(fn)
    wb=openpyxl.load_workbook(path,data_only=True); ws=wb[wb.sheetnames[0]]
    rows=[list(r) for r in ws.iter_rows(values_only=True)]
    # 헤더행 찾기
    hi=next((i for i,r in enumerate(rows) if any('직렬' in str(c) for c in r if c)),None)
    if hi is None: return []
    H=[str(c or '') for c in rows[hi]]
    def idx(kw,*ex):
        for i,c in enumerate(H):
            cc=c.replace('\n','').replace(' ','')
            if kw in cc and not any(e in cc for e in ex): return i
        return -1
    i_jik=idx('직렬'); i_gig=idx('임용'); i_sel=idx('선발'); i_chul=idx('출원')
    i_eung=idx('응시'); i_hap=idx('필기합격','선'); i_cut=idx('합격선','양성'); i_yang=idx('양성')
    out=[]; cur=None
    for r in rows[hi+1:]:
        c=[('' if x is None else x) for x in r]
        def g(i): return c[i] if 0<=i<len(c) else ''
        jik=str(g(i_jik)).strip()
        if jik and '소계' not in jik: cur=parse_jik(jik)
        gig=str(g(i_gig)).strip()
        if not cur or gig in ('소계','총계','계','') or '계' == gig: continue
        cut=num(g(i_cut))
        if cut=='' and num(g(i_yang))=='' and num(g(i_hap))=='': continue
        out.append(["경기","공무원",yr,rd or "1회","공개경쟁",gig,"",cur[0],cur[1],cur[2],cur[3],
                    num(g(i_sel)),num(g(i_chul)),"","",num(g(i_eung)),"","","",num(g(i_hap)),"",cut,"",num(g(i_yang)),"","",fn,""])
    return out

# ---------- 경기 교행 pdf (/ 묶음) ----------
def gyo_pdf(path):
    fn=os.path.basename(path); yr,rd=yr_rd(fn)
    with pdfplumber.open(path) as pdf:
        tabs=[p.extract_table() for p in pdf.pages]
    out=[]
    for t in tabs:
        if not t: continue
        H=[ (c or '').replace('\n','').replace(' ','') for c in t[0]]
        def idx(kw,*ex):
            for i,c in enumerate(H):
                if kw in c and not any(e in c for e in ex): return i
            return -1
        i_jik=idx('직렬'); i_dae=idx('구분') if idx('구분')>=0 else idx('대상')
        # 경기교행: '구분'(시험구분)+직렬(직류)및직급+대상은 별도 열? 헤더 재확인
        i_sel=idx('선발'); i_hap=idx('필기'); i_cut=idx('합격선','양성'); i_yang=idx('양성')
        # 대상열 추정: 직렬 다음 열
        i_dcol=i_jik+1 if i_jik>=0 else -1
        if i_jik<0 or i_sel<0: continue
        for r in t[1:]:
            c=[(x or '') for x in r]
            def g(i): return c[i] if 0<=i<len(c) else ''
            jikraw=g(i_jik).replace('\n','').strip()
            if not jikraw or '직렬' in jikraw: continue
            jr,ju,jg,_=parse_jik(re.sub(r'\s*및\s*직급','',jikraw))
            dcell=g(i_dcol)
            daes=[d for d in re.split(r'[\n/]',dcell) if d.strip()] or ['일반']
            def sp(i):
                v=[x for x in re.split(r'[\n/]',g(i)) if x.strip()!='']
                return v
            sel=sp(i_sel); hap=sp(i_hap); cut=sp(i_cut); yang=sp(i_yang) if i_yang>=0 else []
            N=len(daes)
            def pk(L,k): return L[k] if k<len(L) else (L[-1] if len(L)==1 else '')
            for k in range(N):
                out.append(["경기","교행",yr,rd or "2회","공개경쟁","","",jr,ju,jg or "9급",daenorm(daes[k]),
                            num(pk(sel,k)),"","","","","","","",num(pk(hap,k)),"",num(pk(cut,k)),"",num(pk(yang,k)),"","",fn,""])
    return out

if __name__=='__main__':
    out=[]
    for f in glob.glob(os.path.join(ROOT,'08 경기 공무원','*.xlsx')):
        if '접수' in os.path.basename(f): continue
        r=gong_xlsx(f); print(os.path.basename(f)[:34],'->',len(r)); out+=r
    for f in glob.glob(os.path.join(ROOT,'08 경기 교행','*.pdf')):
        if '붙임2' not in os.path.basename(f) and '합격선' not in os.path.basename(f): continue
        r=gyo_pdf(f); print(os.path.basename(f)[:34],'->',len(r)); out+=r
    OUT=os.path.join(os.path.dirname(__file__),'경기.csv')
    with open(OUT,'w',encoding='utf-8-sig',newline='') as fp:
        w=csv.writer(fp); w.writerow(HEADERS); w.writerows(out)
    print('TOTAL',len(out))
    for r in out[:4]+out[-4:]: print(r[1:12]+[r[21],r[23]])
