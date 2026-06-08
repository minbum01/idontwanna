# -*- coding: utf-8 -*-
"""합격선_통합.xlsx -> 자체완결 HTML 대시보드(라이트테마, 지역별·직렬별 보기). 서버 불필요."""
import openpyxl, json, os
D=os.path.dirname(__file__)
XLSX=os.path.join(D,"합격선_통합.xlsx")
OUT=os.path.join(D,"합격선_관리.html")

wb=openpyxl.load_workbook(XLSX, data_only=True)
def sheet_rows(name):
    ws=wb[name]; H=[c.value for c in ws[1]]
    return [{H[i]:("" if v is None else v) for i,v in enumerate(r)} for r in ws.iter_rows(min_row=2, values_only=True)]
hap = sheet_rows("합격선통합") + sheet_rows("합격선_경력경쟁")
jeop= sheet_rows("2026접수현황")
todo= sheet_rows("작업메모")
data = json.dumps({"hap":hap,"jeop":jeop,"todo":todo}, ensure_ascii=False)

HTML = r"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>합격선 통합 관리</title>
<style>
:root{--bg:#f4f6f9;--card:#fff;--line:#e5e7eb;--txt:#1f2937;--mut:#6b7280;--acc:#2563eb;--acc-bg:#eff6ff;
--cut:#b45309;--ok:#16a34a;--none:#cbd5e1;--head:#f8fafc}
*{box-sizing:border-box}
body{margin:0;font-family:'Pretendard','Malgun Gothic','Segoe UI',sans-serif;background:var(--bg);color:var(--txt);font-size:13px;line-height:1.5}
header{padding:14px 24px;background:#fff;border-bottom:1px solid var(--line);display:flex;align-items:baseline;gap:14px}
header h1{margin:0;font-size:17px;font-weight:700}
.sub{color:var(--mut);font-size:12px}
.tabs{display:flex;gap:2px;padding:0 24px;background:#fff;border-bottom:1px solid var(--line);position:sticky;top:0;z-index:20}
.tab{padding:11px 18px;cursor:pointer;border:none;background:none;color:var(--mut);font-size:14px;border-bottom:2.5px solid transparent}
.tab:hover{color:var(--txt)}
.tab.on{color:var(--acc);border-bottom-color:var(--acc);font-weight:600}
.wrap{padding:20px 24px;max-width:1500px}
.cards{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:18px}
.kpi{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:14px 18px;min-width:120px;box-shadow:0 1px 2px rgba(0,0,0,.03)}
.kpi .n{font-size:24px;font-weight:700;color:var(--acc)}
.kpi .l{color:var(--mut);font-size:12px;margin-top:2px}
.panel{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:16px;box-shadow:0 1px 2px rgba(0,0,0,.03);margin-bottom:16px}
.seg{display:inline-flex;background:#eef1f5;border-radius:10px;padding:3px;margin-bottom:14px}
.seg button{border:none;background:none;padding:7px 16px;border-radius:8px;cursor:pointer;color:var(--mut);font-size:13px;font-weight:600}
.seg button.on{background:#fff;color:var(--acc);box-shadow:0 1px 3px rgba(0,0,0,.1)}
.filters{display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin-bottom:12px}
.fl{display:flex;flex-direction:column;gap:3px}
.fl label{font-size:10px;color:var(--mut);padding-left:2px}
select,input{background:#fff;color:var(--txt);border:1px solid var(--line);border-radius:8px;padding:7px 10px;font-size:13px;outline:none}
select:focus,input:focus{border-color:var(--acc);box-shadow:0 0 0 3px var(--acc-bg)}
input.search{min-width:200px}
.btn{border:1px solid var(--line);background:#fff;border-radius:8px;padding:7px 12px;cursor:pointer;color:var(--mut);font-size:12px}
.btn:hover{background:#f1f5f9}
.count{color:var(--mut);font-size:12px;margin-left:auto;align-self:center}
.tblwrap{overflow:auto;max-height:72vh;border:1px solid var(--line);border-radius:10px;background:#fff}
table{border-collapse:collapse;width:100%}
table.fixed{table-layout:fixed}
table.fixed td{overflow:hidden;text-overflow:ellipsis}
th,td{padding:7px 10px;border-bottom:1px solid #f1f3f5;text-align:left;white-space:nowrap}
thead th{background:var(--head);position:sticky;top:0;cursor:pointer;font-size:12px;font-weight:600;color:#374151;border-bottom:1.5px solid var(--line)}
thead th:hover{color:var(--acc)}
tbody tr:hover{background:#f8fbff}
.r{text-align:right;font-variant-numeric:tabular-nums}
.tag{display:inline-block;padding:1px 8px;border-radius:20px;font-size:11px;font-weight:600}
.t-공무원{background:#dbeafe;color:#1e40af}.t-교행{background:#f3e8ff;color:#6b21a8}
.t-공개경쟁{background:#e0f2fe;color:#075985}.t-경력경쟁{background:#fef3c7;color:#92400e}
.cut{font-weight:700;color:var(--cut);text-align:right;font-variant-numeric:tabular-nums}
.yang{color:#9a3412;font-size:11px}
.matrix th,.matrix td{text-align:center}
.matrix td b{color:var(--acc)}
.have{color:var(--ok);font-weight:700}.no{color:var(--none)}
.pills{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:14px}
.pill{padding:7px 14px;border-radius:20px;border:1px solid var(--line);background:#fff;cursor:pointer;font-size:13px;font-weight:600;color:#374151}
.pill:hover{border-color:var(--acc);color:var(--acc)}
.pill.on{background:var(--acc);color:#fff;border-color:var(--acc)}
.pivot td.v{font-weight:700;color:var(--cut);text-align:center}
.pivot td.e{color:var(--none);text-align:center}
.pivot th{text-align:center}
.note{font-size:12px;color:var(--mut);margin:6px 2px}
.hidden{display:none}
.lvl{border-left:3px solid}
.lv-미반영{border-color:#ef4444}.lv-확인필요{border-color:#f59e0b}.lv-안내{border-color:#3b82f6}.lv-접수현황{border-color:#8b5cf6}
.badge{font-size:11px;padding:2px 8px;border-radius:6px;font-weight:600}
.b-미반영{background:#fee2e2;color:#b91c1c}.b-확인필요{background:#fef3c7;color:#92400e}.b-안내{background:#dbeafe;color:#1e40af}.b-접수현황{background:#ede9fe;color:#6d28d9}
</style></head><body>
<header><h1>📊 합격선 통합 관리</h1><div class="sub" id="meta"></div></header>
<div class="tabs">
 <button class="tab on" data-t="collect">수집현황</button>
 <button class="tab" data-t="hap">합격선 데이터</button>
 <button class="tab" data-t="jeop">2026 접수현황</button>
 <button class="tab" data-t="todo">작업메모</button>
</div>
<div class="wrap">
 <div id="v-collect"></div>
 <div id="v-hap" class="hidden"></div>
 <div id="v-jeop" class="hidden"></div>
 <div id="v-todo" class="hidden"></div>
</div>
<script>
const DB=__DATA__;
const $=s=>document.querySelector(s), $$=s=>[...document.querySelectorAll(s)];
const el=(t,c,h)=>{const e=document.createElement(t);if(c)e.className=c;if(h!=null)e.innerHTML=h;return e};
const uniq=(a,k)=>[...new Set(a.map(r=>r[k]).filter(x=>x!==''&&x!=null))];
const sortKr=a=>a.sort((x,y)=>String(x).localeCompare(String(y),'ko'));
const num=v=>{const n=parseFloat(String(v).replace(/[^0-9.\-]/g,''));return isNaN(n)?null:n};
const regions=sortKr(uniq(DB.hap,'지역'));
$('#meta').textContent=`합격선 ${DB.hap.length.toLocaleString()}행 · ${regions.length}개 지역 · 2026접수 ${DB.jeop.length}행 · 작업메모 ${DB.todo.length}건`;

function tag(v){return `<span class="tag t-${v}">${v}</span>`}
const WIDTHS={'지역':64,'시험종류':74,'연도':58,'회차':52,'시험구분':82,'임용예정기관':110,'직군':84,'직렬':92,'직류':120,'직급':58,'대상':84,'선발예정인원':56,'접수인원':62,'접수여성':56,'경쟁률(접수/선발)':80,'응시인원':56,'응시여성':56,'응시율':62,'경쟁률(응시/선발)':80,'필기합격인원':56,'필기합격여성':56,'합격선':96,'합격선기준':74,'양성평등합격선':70,'최종합격인원':56,'원본파일명':220,'비고':160};
const esc=s=>String(s).replace(/"/g,'&quot;');
function tbl(cols,rows){
 const ww=cols.map(c=>WIDTHS[c.k]||90), total=ww.reduce((a,b)=>a+b,0);
 let h=`<div class="tblwrap"><table class="fixed" style="min-width:${total}px"><colgroup>`+ww.map(w=>`<col style="width:${w}px">`).join('')+'</colgroup><thead><tr>'+cols.map(c=>`<th data-k="${c.k}">${c.t}</th>`).join('')+'</tr></thead><tbody>';
 h+=rows.map(r=>'<tr>'+cols.map(c=>{
   let v=r[c.k]==null?'':r[c.k];
   if(c.tag)return `<td>${v===''?'':tag(v)}</td>`;
   if(c.cut){let y=r['양성평등합격선'];return `<td class="cut" title="${esc(v)}">${v}${y?` <span class="yang">(${y})</span>`:''}</td>`}
   return `<td class="${c.r?'r':''}" title="${esc(v)}">${v}</td>`;
 }).join('')+'</tr>').join('');
 h+='</tbody></table></div>';
 return h;
}
function makeSortable(container,rows,cols,redraw){
 container.querySelectorAll('th').forEach(th=>th.onclick=()=>{
  const k=th.dataset.k; container._dir=(container._sk==k)?-(container._dir||1):1; container._sk=k;
  rows.sort((a,b)=>{const x=num(a[k]),y=num(b[k]); if(x!=null&&y!=null)return (x-y)*container._dir; return String(a[k]).localeCompare(String(b[k]),'ko')*container._dir});
  redraw();
 });
}

/* ============ 수집현황 ============ */
function collect(){
 const v=$('#v-collect');v.innerHTML='';
 const cards=el('div','cards');
 [['지역',regions.length],['합격선 행',DB.hap.length.toLocaleString()],
  ['공무원',DB.hap.filter(r=>r['시험종류']=='공무원').length],['교행',DB.hap.filter(r=>r['시험종류']=='교행').length],
  ['수집/처리 필요',DB.todo.filter(r=>String(r['유형']).startsWith('미반영')||r['유형']=='접수현황 찾기').length]]
  .forEach(([l,n])=>cards.appendChild(el('div','kpi',`<div class="n">${n}</div><div class="l">${l}</div>`)));
 v.appendChild(cards);
 const p=el('div','panel');
 p.appendChild(el('div',null,'<b>지역 × 시험종류 × 연도별 보유 합격선 행수</b>'));
 p.appendChild(el('div','note','숫자=해당 연도 합격선 데이터 행수 · "·"=미수집(작업메모 참고)'));
 const yrs=sortKr(uniq(DB.hap,'연도'));
 let h='<div class="tblwrap"><table class="matrix"><thead><tr><th>지역</th><th>종류</th>'+yrs.map(y=>`<th>${y}</th>`).join('')+'<th>합계</th></tr></thead><tbody>';
 regions.forEach(reg=>['공무원','교행'].forEach(k=>{
  const rs=DB.hap.filter(r=>r['지역']==reg&&r['시험종류']==k);
  const cells=yrs.map(y=>{const c=rs.filter(r=>r['연도']==y).length;return `<td class="${c?'have':'no'}">${c||'·'}</td>`}).join('');
  h+=`<tr><td><b>${reg}</b></td><td>${tag(k)}</td>${cells}<td><b>${rs.length||'·'}</b></td></tr>`;
 }));
 h+='</tbody></table></div>'; p.appendChild(el('div',null,h)); v.appendChild(p);
}

/* ============ 합격선 데이터 (표/지역별/직렬별) ============ */
const HCOLS=[{k:'지역'},{k:'시험종류',tag:1},{k:'연도'},{k:'회차'},{k:'시험구분',tag:1},{k:'임용예정기관',t:'임용기관'},{k:'직렬'},{k:'직류'},{k:'직급'},{k:'대상'},{k:'선발예정인원',t:'선발',r:1},{k:'접수인원',t:'접수',r:1},{k:'경쟁률(접수/선발)',t:'경쟁률'},{k:'응시인원',t:'응시',r:1},{k:'응시율',r:1},{k:'필기합격인원',t:'합격',r:1},{k:'합격선',cut:1},{k:'합격선기준',t:'기준'},{k:'최종합격인원',t:'최종',r:1},{k:'원본파일명',t:'출처'},{k:'비고'}];
HCOLS.forEach(c=>c.t=c.t||c.k);
let hapMode='table';
function hapView(){
 const v=$('#v-hap');v.innerHTML='';
 const seg=el('div','seg');
 [['table','📋 표 보기'],['region','🗺️ 지역별'],['series','🧩 직렬별']].forEach(([m,l])=>{
  const b=el('button',m==hapMode?'on':'',l);b.onclick=()=>{hapMode=m;hapView()};seg.appendChild(b);
 });
 v.appendChild(seg);
 const body=el('div');body.id='hbody';v.appendChild(body);
 ({table:modeTable,region:modeRegion,series:modeSeries})[hapMode]();
}
function modeTable(){
 const b=$('#hbody');b.innerHTML='';
 const p=el('div','panel');const f=el('div','filters');
 const fields=['지역','시험종류','연도','회차','시험구분','직렬','대상'];
 fields.forEach(k=>{
  const w=el('div','fl');w.innerHTML=`<label>${k}</label>`;
  const s=el('select');s.id='f-'+k;s.innerHTML=`<option value="">전체</option>`+sortKr(uniq(DB.hap,k)).map(x=>`<option>${x}</option>`).join('');
  s.onchange=draw;w.appendChild(s);f.appendChild(w);
 });
 const w=el('div','fl');w.innerHTML='<label>검색</label>';const q=el('input','search');q.id='f-q';q.placeholder='직류·기관·파일명…';q.oninput=draw;w.appendChild(q);f.appendChild(w);
 const reset=el('button','btn','초기화');reset.onclick=()=>{fields.forEach(k=>$('#f-'+k).value='');q.value='';draw()};
 const rw=el('div','fl');rw.innerHTML='<label>&nbsp;</label>';rw.appendChild(reset);f.appendChild(rw);
 const cnt=el('span','count');cnt.id='hcnt';f.appendChild(cnt);
 p.appendChild(f);const t=el('div');t.id='htbl';p.appendChild(t);b.appendChild(p);
 function draw(){
  let rows=DB.hap.slice();
  fields.forEach(k=>{const val=$('#f-'+k).value;if(val)rows=rows.filter(r=>String(r[k])==val)});
  const s=$('#f-q').value.trim();if(s)rows=rows.filter(r=>HCOLS.some(c=>String(r[c.k]).includes(s)));
  $('#hcnt').textContent=rows.length.toLocaleString()+'행';
  const render=()=>{$('#htbl').innerHTML=tbl(HCOLS,rows.slice(0,1200));makeSortable($('#htbl'),rows,HCOLS,render);if(rows.length>1200)$('#htbl').insertAdjacentHTML('beforeend',`<p class="note">1200행까지 표시(전체 ${rows.length}). 필터로 좁혀보세요.</p>`)};
  render();
 }
 draw();
}
function modeRegion(){
 const b=$('#hbody');b.innerHTML='';
 const pills=el('div','pills');
 let cur=regions[0];
 regions.forEach(reg=>{const p=el('button','pill'+(reg==cur?' on':''),reg);p.onclick=()=>{cur=reg;[...pills.children].forEach(c=>c.classList.toggle('on',c.textContent==reg));show()};pills.appendChild(p)});
 b.appendChild(pills);
 const panel=el('div','panel');b.appendChild(panel);
 const cols=HCOLS.filter(c=>!['지역','임용예정기관','원본파일명'].includes(c.k));
 function show(){
  const rows=DB.hap.filter(r=>r['지역']==cur).slice().sort((a,b)=>String(a['시험종류']+a['직렬']+a['연도']).localeCompare(String(b['시험종류']+b['직렬']+b['연도']),'ko'));
  const cmu=rows.filter(r=>r['시험종류']=='공무원').length,cgy=rows.filter(r=>r['시험종류']=='교행').length;
  panel.innerHTML=`<div class="cards"><div class="kpi"><div class="n">${rows.length}</div><div class="l">${cur} 합격선 행</div></div>
   <div class="kpi"><div class="n">${cmu}</div><div class="l">공무원</div></div><div class="kpi"><div class="n">${cgy}</div><div class="l">교행</div></div>
   <div class="kpi"><div class="n">${uniq(rows,'직렬').length}</div><div class="l">직렬 수</div></div></div>`;
  const div=el('div');panel.appendChild(div);
  const render=()=>{div.innerHTML=tbl(cols,rows);makeSortable(div,rows,cols,render)};render();
 }
 show();
}
function modeSeries(){
 const b=$('#hbody');b.innerHTML='';
 const p=el('div','panel');
 const f=el('div','filters');
 const mkSel=(id,lbl,opts)=>{const w=el('div','fl');w.innerHTML=`<label>${lbl}</label>`;const s=el('select');s.id=id;s.innerHTML=opts;w.appendChild(s);f.appendChild(w);return s};
 const kindSel=mkSel('s-kind','시험종류',['공무원','교행'].map(x=>`<option>${x}</option>`).join(''));
 const jrSel=mkSel('s-jr','직렬','');
 const juSel=mkSel('s-ju','직류','<option value="">전체</option>');
 const daeSel=mkSel('s-dae','대상','');
 const metricSel=mkSel('s-met','지표',[['합격선','합격선'],['응시율','응시율(%)'],['경쟁률(접수/선발)','경쟁률(접수)'],['경쟁률(응시/선발)','경쟁률(응시)'],['선발예정인원','선발인원'],['응시인원','응시인원']].map(([v,l])=>`<option value="${v}">${l}</option>`).join(''));
 p.appendChild(f);
 const note=el('div','note','지역 × 연도·회차별 비교표(피벗). 같은 시험종류는 기준(과목평균/총점)이 같아 비교 가능합니다.');
 p.appendChild(note);
 const out=el('div');out.id='pv';p.appendChild(out);
 b.appendChild(p);
 function fillJr(){const list=sortKr(uniq(DB.hap.filter(r=>r['시험종류']==kindSel.value),'직렬'));jrSel.innerHTML=list.map(x=>`<option>${x}</option>`).join('')}
 function fillJu(){const list=sortKr(uniq(DB.hap.filter(r=>r['시험종류']==kindSel.value&&r['직렬']==jrSel.value),'직류'));juSel.innerHTML='<option value="">전체</option>'+list.map(x=>`<option>${x}</option>`).join('')}
 function fillDae(){const list=sortKr(uniq(DB.hap.filter(r=>r['시험종류']==kindSel.value&&r['직렬']==jrSel.value),'대상'));daeSel.innerHTML=list.map(x=>`<option ${x=='일반'?'selected':''}>${x}</option>`).join('')}
 kindSel.onchange=()=>{fillJr();fillJu();fillDae();draw()};
 jrSel.onchange=()=>{fillJu();fillDae();draw()};
 juSel.onchange=draw;daeSel.onchange=draw;metricSel.onchange=draw;
 fillJr();fillJu();fillDae();
 function draw(){
  let rows=DB.hap.filter(r=>r['시험종류']==kindSel.value&&r['직렬']==jrSel.value&&r['대상']==daeSel.value);
  if(juSel.value)rows=rows.filter(r=>r['직류']==juSel.value);
  const met=metricSel.value;
  const ycols=sortKr(uniq(rows,'연도')).flatMap(y=>{const rc=sortKr(uniq(rows.filter(r=>r['연도']==y),'회차'));return rc.map(c=>({y,c,key:y+'·'+c}))});
  const regs=sortKr(uniq(rows,'지역'));
  const base=rows[0]?rows[0]['합격선기준']:'';
  const unit=met=='합격선'?(base||'-'):met=='응시율'?'%':met.startsWith('경쟁률')?'배(:1)':'명';
  let h=`<p class="note">지표: <b>${metricSel.options[metricSel.selectedIndex].text}</b> · 단위 <b>${unit}</b> · ${jrSel.value}${juSel.value?'('+juSel.value+')':''} / 대상 ${daeSel.value}</p>`;
  h+='<div class="tblwrap"><table class="pivot"><thead><tr><th>지역</th>'+ycols.map(c=>`<th>${c.key}</th>`).join('')+'</tr></thead><tbody>';
  regs.forEach(reg=>{
   h+=`<tr><td><b>${reg}</b></td>`+ycols.map(c=>{
     const m=rows.find(r=>r['지역']==reg&&r['연도']==c.y&&String(r['회차'])==String(c.c));
     const val=(m&&m[met]!==''&&m[met]!=null)?m[met]:null;
     const y2=(met=='합격선'&&m&&m['양성평등합격선'])?` <span class="yang">(${m['양성평등합격선']})</span>`:'';
     return val!=null?`<td class="v">${val}${y2}</td>`:'<td class="e">·</td>';
   }).join('')+'</tr>';
  });
  h+='</tbody></table></div>';
  if(!regs.length)h='<p class="note">데이터 없음</p>';
  $('#pv').innerHTML=h;
 }
 draw();
}

/* ============ 2026 접수 ============ */
function jeopView(){
 const v=$('#v-jeop');v.innerHTML='';
 const cols=[{k:'지역'},{k:'시험종류',tag:1},{k:'연도'},{k:'회차'},{k:'시험구분',tag:1},{k:'직렬'},{k:'직류'},{k:'직급'},{k:'대상'},{k:'선발예정인원',t:'선발',r:1},{k:'접수인원',t:'접수',r:1},{k:'접수여성',t:'여성',r:1},{k:'경쟁률(접수/선발)',t:'경쟁률'},{k:'비고'}];
 cols.forEach(c=>c.t=c.t||c.k);
 const p=el('div','panel');const f=el('div','filters');
 ['지역','시험종류','직렬'].forEach(k=>{const w=el('div','fl');w.innerHTML=`<label>${k}</label>`;const s=el('select');s.id='j-'+k;s.innerHTML='<option value="">전체</option>'+sortKr(uniq(DB.jeop,k)).map(x=>`<option>${x}</option>`).join('');s.onchange=draw;w.appendChild(s);f.appendChild(w)});
 const cnt=el('span','count');cnt.id='jcnt';f.appendChild(cnt);
 p.appendChild(f);p.appendChild(el('div','note','2026년도 원서접수 단계(합격선 미발표). 경쟁률=접수÷선발 계산.'));
 const t=el('div');t.id='jtbl';p.appendChild(t);v.appendChild(p);
 function draw(){let rows=DB.jeop.slice();['지역','시험종류','직렬'].forEach(k=>{const val=$('#j-'+k).value;if(val)rows=rows.filter(r=>String(r[k])==val)});
  $('#jcnt').textContent=rows.length+'행';const render=()=>{$('#jtbl').innerHTML=tbl(cols,rows);makeSortable($('#jtbl'),rows,cols,render)};render()}
 draw();
}

/* ============ 작업메모 ============ */
function todoView(){
 const v=$('#v-todo');v.innerHTML='';
 const order={'미반영(파싱필요)':0,'접수현황 찾기':1,'확인필요':2,'안내':3};
 const rows=DB.todo.slice().sort((a,b)=>(order[a['유형']]??9)-(order[b['유형']]??9));
 const p=el('div','panel');
 p.appendChild(el('div','note','수집·처리해야 할 항목. 파일을 폴더에 넣고 다시 빌드하면 자동으로 줄어듭니다.'));
 let h='<div class="tblwrap"><table><thead><tr><th>유형</th><th>지역</th><th>종류</th><th>연도</th><th>회차</th><th>항목</th><th>메모</th></tr></thead><tbody>';
 h+=rows.map(r=>{const ty=String(r['유형']);const base=ty.startsWith('미반영')?'미반영':ty=='접수현황 찾기'?'접수현황':ty;
   return `<tr class="lvl lv-${base}"><td><span class="badge b-${base}">${ty}</span></td><td>${r['지역']}</td><td>${r['시험종류']}</td><td>${r['연도']}</td><td>${r['회차']}</td><td>${r['항목']}</td><td style="color:var(--mut)">${r['메모']}</td></tr>`}).join('');
 h+='</tbody></table></div>';p.appendChild(el('div',null,h));v.appendChild(p);
}

collect();hapView();jeopView();todoView();
$$('.tab').forEach(t=>t.onclick=()=>{$$('.tab').forEach(x=>x.classList.remove('on'));t.classList.add('on');
 ['collect','hap','jeop','todo'].forEach(n=>$('#v-'+n).classList.toggle('hidden',n!=t.dataset.t))});
</script></body></html>"""
html=HTML.replace("__DATA__",data)
with open(OUT,"w",encoding="utf-8") as fp: fp.write(html)
print("HTML 생성:",OUT,"|",round(len(html)/1024),"KB | 합격선",len(hap),"접수",len(jeop),"메모",len(todo))
